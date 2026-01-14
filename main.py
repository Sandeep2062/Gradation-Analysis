import openpyxl
import customtkinter as ctk
from tkinter import filedialog, messagebox
import os
import sys
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import winreg
import shutil
from copy import deepcopy
import webbrowser

"""
╔══════════════════════════════════════════════════════════════════╗
║                AGGREGATE GRADATION ANALYZER                     ║
║                                                                   ║
║  Developer: Sandeep (https://github.com/Sandeep2062)            ║
║  Repository: https://github.com/Sandeep2062/Aggregate-Analyzer   ║
║                                                                   ║
║  © 2026 Sandeep - All Rights Reserved                           ║
╚══════════════════════════════════════════════════════════════════╝
"""

# Set appearance
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Resource path for PyInstaller
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Registry Settings
class RegistrySettings:
    def __init__(self):
        self.SOFTWARE_KEY = r"SOFTWARE\AggregateGradationAnalyzer"
        self.app_key = None
        
    def _open_key(self, write=False):
        try:
            if write:
                self.app_key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, self.SOFTWARE_KEY)
            else:
                self.app_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, self.SOFTWARE_KEY, 0, winreg.KEY_READ)
            return True
        except WindowsError:
            try:
                self.app_key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, self.SOFTWARE_KEY)
                return True
            except:
                return False
    
    def _close_key(self):
        if self.app_key:
            winreg.CloseKey(self.app_key)
            self.app_key = None
    
    def save_setting(self, name, value):
        if self._open_key(write=True):
            try:
                if isinstance(value, list):
                    value = "|".join(value)
                winreg.SetValueEx(self.app_key, name, 0, winreg.REG_SZ, str(value))
            except:
                pass
            finally:
                self._close_key()
    
    def load_setting(self, name, default=""):
        if self._open_key():
            try:
                value, _ = winreg.QueryValueEx(self.app_key, name)
                return value
            except WindowsError:
                return default
            finally:
                self._close_key()
        return default
    
    def save_all_settings(self, last_file_path):
        self.save_setting("last_file_path", last_file_path)
    
    def load_all_settings(self):
        last_file_path = self.load_setting("last_file_path", "")
        return {"last_file_path": last_file_path}

registry_settings = RegistrySettings()

def load_workbook_safe(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=False, data_only=False, keep_links=False)
        return wb
    except:
        wb = openpyxl.load_workbook(filepath)
        return wb

def extract_gradation_data(file_path, aggregate_type, log_callback=None):
    try:
        wb = load_workbook_safe(file_path)
        
        # Determine which sheet to use based on aggregate type
        sheet_name = None
        if aggregate_type.lower() == "fine aggregate":
            # Look for gradation analysis sheet in fine aggregate file
            for name in wb.sheetnames:
                if "gradation" in name.lower():
                    sheet_name = name
                    break
        elif aggregate_type.lower() == "sub-base":
            # Look for gradation sheet in sub-base file
            for name in wb.sheetnames:
                if "gradation" in name.lower():
                    sheet_name = name
                    break
        elif aggregate_type.lower() == "coarse aggregate":
            # Look for gradation analysis sheet in coarse aggregate file
            for name in wb.sheetnames:
                if "gradation" in name.lower():
                    sheet_name = name
                    break
        
        if not sheet_name:
            if log_callback:
                log_callback(f"⚠ No gradation sheet found in {os.path.basename(file_path)}")
            return None, None, None, None
        
        ws = wb[sheet_name]
        
        # Extract sieve sizes and passing percentages
        sieve_sizes = []
        passing_percentages = []
        weight_retained = []
        
        # Find the header row with sieve sizes
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "sieve" in str(cell_value).lower():
                header_row = row
                break
        
        if not header_row:
            if log_callback:
                log_callback(f"⚠ Could not find sieve data in {os.path.basename(file_path)}")
            return None, None, None, None
        
        # Extract data
        data_row = header_row + 1
        while True:
            sieve_size = ws.cell(row=data_row, column=1).value
            if not sieve_size:
                break
                
            # Skip non-numeric sieve sizes
            try:
                sieve_size = float(sieve_size)
            except:
                data_row += 1
                continue
                
            # Get passing percentage
            passing_pct = ws.cell(row=data_row, column=5).value  # Column 5 typically has % passing
            if passing_pct is None:
                # Try other columns if column 5 doesn't have data
                for col in range(6, 10):
                    passing_pct = ws.cell(row=data_row, column=col).value
                    if passing_pct is not None:
                        break
            
            # Get weight retained
            weight = ws.cell(row=data_row, column=2).value  # Column 2 typically has weight retained
            
            if passing_pct is not None:
                try:
                    sieve_sizes.append(sieve_size)
                    passing_percentages.append(float(passing_pct))
                    if weight is not None:
                        weight_retained.append(float(weight))
                    else:
                        weight_retained.append(0)
                except:
                    pass
            
            data_row += 1
        
        # Extract limits if available
        min_limits = []
        max_limits = []
        
        # Try to find limits in the same row as the sieve sizes
        limits_row = header_row
        for col in range(6, 10):  # Check columns after passing percentage
            cell_value = ws.cell(row=limits_row, column=col).value
            if cell_value and ("min" in str(cell_value).lower() or "max" in str(cell_value).lower()):
                # This might be the limits column
                min_col = col
                max_col = col + 1
                
                # Extract limits for each sieve size
                for i, row in enumerate(range(header_row + 1, header_row + 1 + len(sieve_sizes))):
                    min_val = ws.cell(row=row, column=min_col).value
                    max_val = ws.cell(row=row, column=max_col).value
                    
                    if min_val is not None:
                        try:
                            min_limits.append(float(min_val))
                        except:
                            min_limits.append(0)
                    else:
                        min_limits.append(0)
                        
                    if max_val is not None:
                        try:
                            max_limits.append(float(max_val))
                        except:
                            max_limits.append(100)
                    else:
                        max_limits.append(100)
                
                break
        
        wb.close()
        
        if log_callback:
            log_callback(f"✓ Loaded {len(sieve_sizes)} sieve sizes from {os.path.basename(file_path)}")
        
        return sieve_sizes, passing_percentages, weight_retained, (min_limits, max_limits)
        
    except Exception as e:
        if log_callback:
            log_callback(f"✖ Error loading {os.path.basename(file_path)}: {e}")
        return None, None, None, None

class AggregateGradationAnalyzer:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("Aggregate Gradation Analyzer")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 700)
        
        try:
            self.root.iconbitmap(resource_path("icon.ico"))
        except:
            pass
        
        settings = registry_settings.load_all_settings()
        
        self.file_path = ctk.StringVar(value=settings.get("last_file_path", ""))
        self.aggregate_type = ctk.StringVar(value="Fine Aggregate")
        
        # Data variables
        self.sieve_sizes = []
        self.passing_percentages = []
        self.weight_retained = []
        self.min_limits = []
        self.max_limits = []
        
        self.setup_ui()
        
    def setup_ui(self):
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(0, weight=1)
        
        self.create_sidebar()
        self.create_main_content()
        
    def create_sidebar(self):
        self.sidebar = ctk.CTkFrame(self.root, width=260, corner_radius=0)
        self.sidebar.grid(row=0, column=0, rowspan=2, sticky="nsew")
        self.sidebar.grid_rowconfigure(10, weight=1)
        
        # Logo
        logo_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        logo_frame.grid(row=0, column=0, padx=20, pady=(30, 10))
        
        try:
            from PIL import Image
            logo_img = Image.open(resource_path("logo.png"))
            logo_img = logo_img.resize((70, 70), Image.Resampling.LANCZOS)
            logo_photo = ctk.CTkImage(light_image=logo_img, dark_image=logo_img, size=(70, 70))
            logo_label = ctk.CTkLabel(logo_frame, image=logo_photo, text="")
            logo_label.pack()
        except:
            logo_label = ctk.CTkLabel(logo_frame, text="📊", font=ctk.CTkFont(size=50))
            logo_label.pack()
        
        title_label = ctk.CTkLabel(self.sidebar, text="AGGREGATE\nGRADATION ANALYZER", 
                                   font=ctk.CTkFont(size=22, weight="bold"))
        title_label.grid(row=1, column=0, padx=20, pady=(0, 30))
        
        # Aggregate type selection
        type_label = ctk.CTkLabel(self.sidebar, text="📋 Aggregate Type", 
                                  font=ctk.CTkFont(size=15, weight="bold"), anchor="w")
        type_label.grid(row=2, column=0, padx=20, pady=(10, 15), sticky="ew")
        
        self.fine_radio = ctk.CTkRadioButton(self.sidebar, text="Fine Aggregate", 
                                            variable=self.aggregate_type, value="Fine Aggregate",
                                            command=self.update_ui,
                                            font=ctk.CTkFont(size=13))
        self.fine_radio.grid(row=3, column=0, padx=30, pady=8, sticky="w")
        
        self.coarse_radio = ctk.CTkRadioButton(self.sidebar, text="Coarse Aggregate", 
                                             variable=self.aggregate_type, value="Coarse Aggregate",
                                             command=self.update_ui,
                                             font=ctk.CTkFont(size=13))
        self.coarse_radio.grid(row=4, column=0, padx=30, pady=8, sticky="w")
        
        self.subbase_radio = ctk.CTkRadioButton(self.sidebar, text="Sub-Base", 
                                              variable=self.aggregate_type, value="Sub-Base",
                                              command=self.update_ui,
                                              font=ctk.CTkFont(size=13))
        self.subbase_radio.grid(row=5, column=0, padx=30, pady=8, sticky="w")
        
        # File selection
        file_label = ctk.CTkLabel(self.sidebar, text="📁 Excel File", 
                                 font=ctk.CTkFont(size=15, weight="bold"), anchor="w")
        file_label.grid(row=6, column=0, padx=20, pady=(25, 10), sticky="ew")
        
        self.file_entry = ctk.CTkEntry(self.sidebar, textvariable=self.file_path, 
                                      placeholder_text="Select Excel file...",
                                      height=40, font=ctk.CTkFont(size=11))
        self.file_entry.grid(row=7, column=0, padx=20, pady=(0, 10), sticky="ew")
        
        file_btn = ctk.CTkButton(self.sidebar, text="Browse", 
                                command=self.pick_file, width=220, height=40,
                                font=ctk.CTkFont(size=12, weight="bold"))
        file_btn.grid(row=8, column=0, padx=20, pady=(0, 20))
        
        # Data manipulation
        data_label = ctk.CTkLabel(self.sidebar, text="🔧 Data Manipulation", 
                                 font=ctk.CTkFont(size=15, weight="bold"), anchor="w")
        data_label.grid(row=9, column=0, padx=20, pady=(10, 15), sticky="ew")
        
        random_btn = ctk.CTkButton(self.sidebar, text="🎲 Generate Random", 
                                  command=self.generate_random_data, width=220, height=40,
                                  font=ctk.CTkFont(size=12, weight="bold"))
        random_btn.grid(row=10, column=0, padx=20, pady=5)
        
        copy_btn = ctk.CTkButton(self.sidebar, text="📋 Copy Weight Retained", 
                                command=self.copy_weight_retained, width=220, height=40,
                                font=ctk.CTkFont(size=12, weight="bold"))
        copy_btn.grid(row=11, column=0, padx=20, pady=5)
        
        # Social links
        social_label = ctk.CTkLabel(self.sidebar, text="🔗 Connect", 
                                    font=ctk.CTkFont(size=13, weight="bold"), anchor="w")
        social_label.grid(row=12, column=0, padx=20, pady=(20, 10), sticky="ew")
        
        github_btn = ctk.CTkButton(self.sidebar, text="🐙 GitHub", 
                                   command=lambda: webbrowser.open("https://github.com/Sandeep2062/Aggregate-Analyzer"),
                                   width=220, height=35, font=ctk.CTkFont(size=12, weight="bold"))
        github_btn.grid(row=13, column=0, padx=20, pady=5)
        
        insta_btn = ctk.CTkButton(self.sidebar, text="📷 Instagram", 
                                  command=lambda: webbrowser.open("https://www.instagram.com/sandeep._.2062/"),
                                  width=220, height=35, font=ctk.CTkFont(size=12, weight="bold"),
                                  fg_color="#E1306C", hover_color="#C13584")
        insta_btn.grid(row=14, column=0, padx=20, pady=(5, 30))
        
    def create_main_content(self):
        self.main_frame = ctk.CTkFrame(self.root)
        self.main_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(1, weight=1)
        
        # Create matplotlib figure for the graph
        self.fig = Figure(figsize=(8, 6), dpi=100)
        self.ax = self.fig.add_subplot(111)
        
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.main_frame)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        # Create data table
        self.table_frame = ctk.CTkFrame(self.main_frame)
        self.table_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        
        # Create table headers
        headers = ["Sieve Size (mm)", "Weight Retained (g)", "Passing (%)"]
        self.table_entries = []
        
        for i, header in enumerate(headers):
            label = ctk.CTkLabel(self.table_frame, text=header, font=ctk.CTkFont(size=12, weight="bold"))
            label.grid(row=0, column=i, padx=10, pady=5)
        
        # Create table rows (initially empty)
        for i in range(10):  # Start with 10 empty rows
            row_entries = []
            for j in range(3):
                entry = ctk.CTkEntry(self.table_frame, width=120)
                entry.grid(row=i+1, column=j, padx=5, pady=2)
                row_entries.append(entry)
            self.table_entries.append(row_entries)
        
        # Add scrollbar frame
        self.scroll_frame = ctk.CTkScrollableFrame(self.table_frame, height=200)
        self.scroll_frame.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=10, pady=5)
        
        # Footer
        footer_label = ctk.CTkLabel(self.main_frame, 
                                    text="© 2026 Sandeep | github.com/Sandeep2062/Aggregate-Analyzer", 
                                    font=ctk.CTkFont(size=11), text_color="gray60")
        footer_label.grid(row=2, column=0, pady=(5, 10))
        
    def update_ui(self):
        # Update UI based on selected aggregate type
        if self.file_path.get():
            self.load_data()
        
    def pick_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if path:
            self.file_path.set(path)
            registry_settings.save_all_settings(path)
            self.load_data()
        
    def load_data(self):
        if not self.file_path.get():
            return
            
        self.sieve_sizes, self.passing_percentages, self.weight_retained, limits = extract_gradation_data(
            self.file_path.get(), 
            self.aggregate_type.get(),
            log_callback=self.log
        )
        
        if limits:
            self.min_limits, self.max_limits = limits
        
        self.update_graph()
        self.update_table()
        
    def update_graph(self):
        self.ax.clear()
        
        if not self.sieve_sizes or not self.passing_percentages:
            self.ax.set_title("No data to display")
            self.canvas.draw()
            return
        
        # Plot the gradation curve
        self.ax.plot(self.sieve_sizes, self.passing_percentages, 'bo-', label='Actual')
        
        # Plot limits if available
        if self.min_limits and self.max_limits:
            self.ax.fill_between(self.sieve_sizes, self.min_limits, self.max_limits, 
                                color='gray', alpha=0.3, label='Specification Limits')
        
        self.ax.set_xlabel('Sieve Size (mm)')
        self.ax.set_ylabel('Passing (%)')
        self.ax.set_title(f'{self.aggregate_type.get()} Gradation Analysis')
        self.ax.grid(True)
        self.ax.legend()
        self.ax.invert_xaxis()  # Larger sieve sizes on the left
        
        self.canvas.draw()
        
    def update_table(self):
        # Clear existing entries in the scroll frame
        for widget in self.scroll_frame.winfo_children():
            widget.destroy()
        
        # Create new entries based on data
        self.table_entries = []
        
        for i in range(len(self.sieve_sizes)):
            row_entries = []
            for j in range(3):
                entry = ctk.CTkEntry(self.scroll_frame, width=120)
                entry.grid(row=i, column=j, padx=5, pady=2)
                
                # Set initial values
                if j == 0:  # Sieve size
                    entry.insert(0, str(self.sieve_sizes[i]))
                    entry.configure(state="disabled")  # Make sieve size read-only
                elif j == 1:  # Weight retained
                    entry.insert(0, str(self.weight_retained[i]))
                    # Bind event to update graph when value changes
                    entry.bind("<KeyRelease>", lambda e, idx=i: self.update_from_table(idx))
                elif j == 2:  # Passing percentage
                    entry.insert(0, str(self.passing_percentages[i]))
                    # Bind event to update graph when value changes
                    entry.bind("<KeyRelease>", lambda e, idx=i: self.update_from_table(idx))
                
                row_entries.append(entry)
            self.table_entries.append(row_entries)
    
    def update_from_table(self, idx):
        try:
            # Update weight retained
            weight = float(self.table_entries[idx][1].get())
            self.weight_retained[idx] = weight
            
            # Update passing percentage
            passing = float(self.table_entries[idx][2].get())
            self.passing_percentages[idx] = passing
            
            # Update graph
            self.update_graph()
        except:
            pass  # Ignore invalid input
    
    def generate_random_data(self):
        if not self.sieve_sizes:
            messagebox.showwarning("Warning", "Please load data first")
            return
        
        # Generate random passing percentages within limits if available
        for i in range(len(self.sieve_sizes)):
            if self.min_limits and self.max_limits and i < len(self.min_limits) and i < len(self.max_limits):
                # Generate random value within limits
                min_val = self.min_limits[i]
                max_val = self.max_limits[i]
                random_val = np.random.uniform(min_val, max_val)
            else:
                # Generate random value between 0 and 100
                random_val = np.random.uniform(0, 100)
            
            self.passing_percentages[i] = random_val
            
            # Update table
            self.table_entries[i][2].delete(0, "end")
            self.table_entries[i][2].insert(0, f"{random_val:.2f}")
        
        # Update graph
        self.update_graph()
        
        self.log("🎲 Generated random passing percentages")
    
    def copy_weight_retained(self):
        if not self.weight_retained:
            messagebox.showwarning("Warning", "No weight retained data to copy")
            return
        
        # Create a string with weight retained values
        weight_str = "\n".join([f"{w:.2f}" for w in self.weight_retained])
        
        # Copy to clipboard
        self.root.clipboard_clear()
        self.root.clipboard_append(weight_str)
        
        messagebox.showinfo("Success", "Weight retained values copied to clipboard")
        self.log("📋 Copied weight retained values to clipboard")
    
    def log(self, message):
        # Simple log function (could be expanded to show in a log window)
        print(message)
        
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = AggregateGradationAnalyzer()
    app.run()